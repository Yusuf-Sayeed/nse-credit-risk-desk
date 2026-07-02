#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from merton_model import merton_df
from expected_loss import el_df
from stress_testing import stress_df


# In[2]:


portfolio_losses = np.load("data/portfolio_losses.npy")


# In[3]:


wb = Workbook()

# Rename default sheet
ws_summary = wb.active
ws_summary.title = "Summary"

# Create remaining sheets
ws_merton = wb.create_sheet("Merton Results")
ws_el = wb.create_sheet("Expected Loss")
ws_stress = wb.create_sheet("Stress Testing")
ws_loss = wb.create_sheet("Loss Distribution")


# In[4]:


ws_summary['A1'] = 'NSE Credit Risk Dashboard'
ws_summary['A2'] = 'Portfolio: 10 Indian Listed Companies | Model: Merton Structural + Monte Carlo'
ws_summary['A4'] = 'Key Metrics'


# In[5]:


ws_summary['A5'] = 'Portfolio Size'
ws_summary['B5'] = '10 Companies'

ws_summary['A6'] = 'Total Exposure'
ws_summary['B6'] = el_df['EAD'].sum()

ws_summary['A7'] = 'Expected Loss'
ws_summary['B7'] = '=SUM(\'Expected Loss\'!E2:E11)'

# Highest risk company by PD (Merton model)
highest_risk_company = merton_df["PD"].idxmax()
highest_risk_pd = merton_df["PD"].max()

# Most dangerous stress scenario by Credit VaR (unexpected loss)
worst_scenario = stress_df["Credit VaR"].idxmax()
worst_scenario_var = stress_df.loc[worst_scenario, "Credit VaR"]

# Credit VaR 99% — use baseline VaR from stress_df, not a guess
baseline_var = stress_df.loc["Baseline", "VaR 99%"]

ws_summary['A8'] = 'Credit VaR (99%)'
ws_summary['B8'] = baseline_var

ws_summary['A9'] = 'Highest Risk Company (by PD)'
ws_summary['B9'] = f"{highest_risk_company} ({highest_risk_pd:.2%})"

ws_summary['A10'] = 'Most Dangerous Stress Scenario'
ws_summary['B10'] = f"{worst_scenario} (₹{worst_scenario_var:,.0f} Credit VaR)"


# In[6]:


# Title formatting
ws_summary['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_summary['A1'].fill = PatternFill('solid', start_color='1F4E79')
ws_summary['A2'].font = Font(italic=True, size=10, color='595959')

# Key Metrics header
ws_summary['A4'].font = Font(bold=True, size=12, color='FFFFFF')
ws_summary['A4'].fill = PatternFill('solid', start_color='2E75B6')

# Column widths
ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 25
ws_summary.merge_cells('A1:B1')
ws_summary['A1'].alignment = Alignment(horizontal='center')
ws_summary.merge_cells('A4:B4')

# Format currency cells
for row in [6, 7, 8]:
    ws_summary[f'B{row}'].number_format = '₹#,##0'


# In[7]:


headers = ['Company', 'Asset Value (V)', 'Asset Volatility', 'Distance to Default', 'Probability of Default']

for col, header in enumerate(headers, 1):
    cell = ws_merton.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E79')
    cell.alignment = Alignment(horizontal='center')


# In[8]:


for row, (name, data) in enumerate(merton_df.iterrows(), 2):
    ws_merton.cell(row=row, column=1).value = name
    ws_merton.cell(row=row, column=2).value = data['V']
    ws_merton.cell(row=row, column=3).value = data['sigma_V']
    ws_merton.cell(row=row, column=4).value = data['DD']
    ws_merton.cell(row=row, column=5).value = data['PD']

    # Format numbers
    ws_merton.cell(row=row, column=2).number_format = '₹#,##0'
    ws_merton.cell(row=row, column=3).number_format = '0.0%'
    ws_merton.cell(row=row, column=4).number_format = '0.00'
    ws_merton.cell(row=row, column=5).number_format = '0.0000%'

# Column widths
for col, width in zip(['A','B','C','D','E'], [25, 20, 18, 20, 22]):
    ws_merton.column_dimensions[col].width = width


# In[9]:


headers = ['Company', 'PD', 'LGD', 'EAD', 'Expected Loss']

for col, header in enumerate(headers, 1):
    cell = ws_el.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E79')
    cell.alignment = Alignment(horizontal='center')


# In[10]:


for row, (name, data) in enumerate(el_df.iterrows(), 2):
    ws_el.cell(row=row, column=1).value = name
    ws_el.cell(row=row, column=2).value = data['PD']
    ws_el.cell(row=row, column=3).value = data['LGD']
    ws_el.cell(row=row, column=4).value = data['EAD']
    ws_el.cell(row=row, column=5).value = data['Expected Loss']

    # Format numbers
    ws_el.cell(row=row, column=2).number_format = '0.0000%'
    ws_el.cell(row=row, column=3).number_format = '0.0%'
    ws_el.cell(row=row, column=4).number_format = '₹#,##0'
    ws_el.cell(row=row, column=5).number_format = '₹#,##0'

# Column widths
for col, width in zip(['A','B','C','D','E'], [25, 15, 10, 20, 20]):
    ws_el.column_dimensions[col].width = width


# In[11]:


headers = ['Scenario', 'Expected Loss', 'Credit VaR 99%', 'Unexpected Loss']

for col, header in enumerate(headers, 1):
    cell = ws_stress.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E79')
    cell.alignment = Alignment(horizontal='center')


# In[12]:


for row, (scenario, data) in enumerate(stress_df.iterrows(), 2):
    ws_stress.cell(row=row, column=1).value = scenario
    ws_stress.cell(row=row, column=2).value = data['EL']
    ws_stress.cell(row=row, column=3).value = data['VaR 99%']
    ws_stress.cell(row=row, column=4).value = data['Credit VaR']

    # Format numbers
    for col in [2, 3, 4]:
        ws_stress.cell(row=row, column=col).number_format = '₹#,##0'

# Highlight worst scenario (highest Credit VaR) in red
worst_row_idx = stress_df["Credit VaR"].values.argmax() + 2  # +2: row 1 is header, data starts at row 2

for col in [1, 2, 3, 4]:
    ws_stress.cell(row=worst_row_idx, column=col).font = Font(bold=True, color='FF0000')

# Column widths
for col, width in zip(['A','B','C','D'], [20, 20, 20, 20]):
    ws_stress.column_dimensions[col].width = width


# In[13]:


# Header
ws_loss.cell(row=1, column=1).value = 'Loss Bucket (Lower Bound)'
ws_loss.cell(row=1, column=2).value = 'Frequency'

for col in [1, 2]:
    cell = ws_loss.cell(row=1, column=col)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E79')
    cell.alignment = Alignment(horizontal='center')

# Compute histogram
counts, bin_edges = np.histogram(portfolio_losses, bins=50)
bin_starts = bin_edges[:-1]  # drop the final right edge — one start per bin, aligned with counts

# Write Data
for row, (count, start) in enumerate(zip(counts, bin_starts), 2):
    ws_loss.cell(row=row, column=1).value = round(start, 0)
    ws_loss.cell(row=row, column=2).value = int(count)
    ws_loss.cell(row=row, column=1).number_format = '₹#,##0'

top_pd_company = merton_df["PD"].idxmax()
ws_loss['D1'] = f'Note: High zero-loss frequency reflects low PD portfolio. {top_pd_company} carries the highest standalone default probability.'

# Column widths
ws_loss.column_dimensions['A'].width = 20
ws_loss.column_dimensions['B'].width = 15


# In[14]:


os.makedirs('outputs', exist_ok=True)
wb.save('outputs/NSE_Credit_Risk_Desk.xlsx')
print("Dashboard saved successfully")


# In[ ]:




